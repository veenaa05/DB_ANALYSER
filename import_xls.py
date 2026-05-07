"""
Import Onex_QueryDetails.xlsx into PostgreSQL dmv_snapshots table.

Sheet mapping:
  Query1  → index_fragmentation   (index fragmentation)
  Query2  → missing_indexes       (missing index recommendations)
  Query3  → unused_indexes        (unused / underused indexes)
  Query4  → slow_queries          (top CPU/elapsed queries)
  Query5  → deadlocks             (XML deadlock reports)
  Query6  → table_sizes           (table disk usage)
  Query7  → server_info           (server hardware/config)
  Details → db_info               (database sizes)
  Jobs    → sql_agent_jobs        (SQL Agent job history)
"""

import json
import psycopg2
import openpyxl
from datetime import datetime, date, time

XLS = r"C:\Users\suraj.dubey\Desktop\Onex_QueryDetails.xlsx"
RUN_ID = None   # will be resolved below

def safe(v):
    """Convert value to JSON-safe type."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return str(v)
    if isinstance(v, time):
        return str(v)
    if isinstance(v, float) and v != v:   # NaN
        return None
    return v

def sheet_to_records(ws, skip_empty=True):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        if skip_empty and all(v is None for v in row):
            continue
        records.append({headers[i]: safe(v) for i, v in enumerate(row)})
    return headers, records

def insert_dmv(cur, run_id, dmv_type, records):
    data_json = json.dumps(records, default=str)
    cur.execute("""
        INSERT INTO dmv_snapshots (run_id, dmv_type, data_json, row_count)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """, (run_id, dmv_type, data_json, len(records)))
    print(f"  OK {dmv_type}: {len(records)} rows")

# ── Connect ───────────────────────────────────────────────────────────────────
conn = psycopg2.connect(host='192.168.202.135', port=5432, database='dbanalyser',
                        user='dbanalyser_user', password='dbanalyser_user')
conn.autocommit = True
cur = conn.cursor()

# Get latest run id
cur.execute("SELECT id FROM runs ORDER BY id DESC LIMIT 1")
row = cur.fetchone()
if not row:
    print("No runs found — seed the DB first")
    exit(1)
RUN_ID = row[0]
print(f"Inserting DMV data for run_id={RUN_ID}\n")

# ── Load workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLS, read_only=True, data_only=True)

# ── Details: DB sizes ─────────────────────────────────────────────────────────
ws = wb["Details"]
rows = list(ws.iter_rows(values_only=True))
db_info = []
for r in rows[1:]:
    if r[0] and r[1] and str(r[0]).strip() not in ("Point1", "DB Name", ""):
        db_info.append({"db_name": str(r[0]).strip(), "size": str(r[1]).strip()})
if db_info:
    insert_dmv(cur, RUN_ID, "db_info", db_info)

# ── Jobs Details: SQL Agent jobs ──────────────────────────────────────────────
_, jobs = sheet_to_records(wb["Jobs Details"])
if jobs:
    insert_dmv(cur, RUN_ID, "sql_agent_jobs", jobs)

# ── Query1: Index fragmentation ───────────────────────────────────────────────
_, q1 = sheet_to_records(wb["Query1"])
# Rename columns to match dashboard expectations
frag_records = []
for r in q1:
    frag_records.append({
        "schema_name":                  r.get("Schema"),
        "table_name":                   r.get("Table"),
        "index_name":                   r.get("Index"),
        "avg_fragmentation_pct":        r.get("avg_fragmentation_in_percent"),
        "page_count":                   r.get("page_count"),
    })
insert_dmv(cur, RUN_ID, "index_fragmentation", frag_records)

# ── Query2: Missing indexes ───────────────────────────────────────────────────
_, q2 = sheet_to_records(wb["Query2"])
missing_idx = []
for r in q2:
    missing_idx.append({
        "database_id":          r.get("DatabaseID"),
        "improvement_measure":  r.get("Avg_Estimated_Impact"),
        "last_user_seek":       r.get("Last_User_Seek"),
        "table_name":           r.get("TableName"),
        "create_statement":     r.get("Create_Statement"),
    })
insert_dmv(cur, RUN_ID, "dmv_missing_indexes", missing_idx)

# ── Query3: Unused / underused indexes ───────────────────────────────────────
_, q3 = sheet_to_records(wb["Query3"])
unused_idx = []
for r in q3:
    user_seeks   = r.get("UserSeek") or 0
    user_scans   = r.get("UserScans") or 0
    user_lookups = r.get("UserLookups") or 0
    total_reads  = (user_seeks or 0) + (user_scans or 0) + (user_lookups or 0)
    flag = "UNUSED" if total_reads == 0 else ("LOW_USE" if total_reads < 100 else "ACTIVE")
    unused_idx.append({
        "object_name":   r.get("ObjectName"),
        "index_name":    r.get("IndexName"),
        "index_id":      r.get("IndexID"),
        "user_seeks":    user_seeks,
        "user_scans":    user_scans,
        "user_lookups":  user_lookups,
        "user_updates":  r.get("UserUpdates"),
        "table_rows":    r.get("TableRows"),
        "drop_stmt":     r.get("drop statement"),
        "flag":          flag,
    })
insert_dmv(cur, RUN_ID, "dmv_index_usage", unused_idx)

# ── Query4: Top slow / high CPU queries ───────────────────────────────────────
_, q4 = sheet_to_records(wb["Query4"])
slow_q = []
for r in q4:
    query_text = str(r.get("Query") or "")
    # Extract object name from CREATE PROCEDURE / FUNCTION line
    obj_name = ""
    import re
    m = re.search(r'CREATE\s+(?:PROCEDURE|FUNCTION|VIEW)\s+\[?(?:\w+)\]?\.\[?(\w+)\]?',
                  query_text, re.IGNORECASE)
    if m:
        obj_name = m.group(1)
    slow_q.append({
        "object_name":              obj_name or query_text[:80],
        "query_text":               query_text[:500],
        "execution_count":          r.get("execution_count"),
        "total_cpu":                r.get("Total_CPU"),
        "avg_cpu_ms":               r.get("average_CPU_inSeconds"),
        "total_elapsed_time":       r.get("total_elapsed_time"),
        "avg_elapsed_ms":           r.get("total_elapsed_time_inSeconds"),
    })
insert_dmv(cur, RUN_ID, "dmv_slow_queries", slow_q)

# ── Query5: Deadlocks ─────────────────────────────────────────────────────────
_, q5 = sheet_to_records(wb["Query5"])
deadlocks = []
for r in q5:
    if r.get("Utc time") or r.get("Local time"):
        deadlocks.append({
            "utc_time":    r.get("Utc time"),
            "local_time":  r.get("Local time"),
            "xml_report":  str(r.get("XML Deadlock Report") or "")[:2000],
        })
if deadlocks:
    insert_dmv(cur, RUN_ID, "dmv_deadlocks", deadlocks)

# ── Query6: Table sizes ───────────────────────────────────────────────────────
_, q6 = sheet_to_records(wb["Query6"])
table_sizes = []
for r in q6:
    table_sizes.append({
        "table_name":     r.get("TableName"),
        "index_name":     r.get("indexName"),
        "row_counts":     r.get("RowCounts"),
        "total_pages":    r.get("TotalPages"),
        "used_pages":     r.get("UsedPages"),
        "data_pages":     r.get("DataPages"),
        "total_size_mb":  r.get("TotalSpaceMB"),
        "used_size_mb":   r.get("UsedSpaceMB"),
        "data_size_mb":   r.get("DataSpaceMB"),
    })
insert_dmv(cur, RUN_ID, "dmv_table_sizes", table_sizes)

# ── Query7: Server info ───────────────────────────────────────────────────────
_, q7 = sheet_to_records(wb["Query7"])
if q7:
    insert_dmv(cur, RUN_ID, "server_info", q7)

wb.close()
conn.close()
print("\nAll done! DMV data imported successfully.")
print("Refresh the dashboard to see Live DB Analysis pages populated.")
